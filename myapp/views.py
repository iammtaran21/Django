from django.shortcuts import render
from openpyxl import load_workbook
from django.http import HttpResponse
from .models import Task
import pandas as pd
# Create your views here.
def index(request):
    return HttpResponse("Hello")

def task_list(request):
    data = pd.read_excel('/Users/taransingh/Downloads/tracker_excel.xlsx', sheet_name='Sheet1')
    tasks = []
    for index, row in data.iterrows():
        task_name = row['task_name']
        assigned_to = row['assigned_to']
        tasks.append({'task_name': task_name, 'assigned_to': assigned_to})
    return render(request, 'task_list.html', {'tasks': tasks})
    # if request.method == 'POST' and request.FILES['file']:
    #     file = request.FILES['file']
    #     wb = load_workbook(filename=file, read_only=True)
    #     ws = wb.active
    #     for row in ws.iter_rows(min_row=2, values_only=True):
    #         task_name, assigned_to = row
    #         task = Task.objects.create(task_name=task_name, assigned_to=assigned_to)
    #         tasks.append(task)
    # else:
    #     tasks = Task.objects.all()
    # return render(request, 'myapp/task_list.html', {'tasks': tasks})
